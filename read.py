import datetime
import openpyxl

def read_meal(meal):
    #considerar meal = 1 para almoço e meal = 2 para janta
    path = ('./cardapio.xlsx')
    wb = openpyxl.load_workbook(path)
    sheet = wb.active  
    day = datetime.datetime.today().weekday()
    proteina = sheet.cell(column=meal*2, row=day+3).value
    massa = sheet.cell(column=meal*2+1, row=day+3).value
    
    return [proteina, massa]
    
    
    
    